"""Report export service for PDF, Excel, and CSV formats."""
import io
import csv
from typing import Dict, Any, List
from datetime import datetime

# Handle optional imports gracefully
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ReportExportService:
    """报表导出服务 - Report export service supporting CSV, Excel, and PDF formats."""

    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = "report.csv") -> bytes:
        """Export data to CSV format with UTF-8 BOM.

        Args:
            data: List of dictionaries containing report data.
            filename: Optional filename (for reference purposes).

        Returns:
            CSV content as UTF-8 encoded bytes with BOM for Excel compatibility.
        """
        if not data:
            return b""

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        # UTF-8 BOM for Excel compatibility
        return output.getvalue().encode('utf-8-sig')

    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        sheet_name: str = "Report",
        title: str = "Report"
    ) -> bytes:
        """Export data to Excel format using openpyxl with formatting.

        Args:
            data: List of dictionaries containing report data.
            sheet_name: Name of the worksheet.
            title: Title to display at the top of the sheet.

        Returns:
            Excel file content as bytes.

        Raises:
            ImportError: If openpyxl is not installed.
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not installed. Install with: pip install openpyxl==3.1.0")

        if not data:
            return b""

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Add title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Add timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Add headers with styling
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add data rows
        for row_idx, row_data in enumerate(data, 5):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(key)
                # Handle None values
                if value is None:
                    cell.value = ""
                else:
                    cell.value = value

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_to_pdf(
        data: List[Dict[str, Any]],
        title: str = "Report",
        subtitle: str = ""
    ) -> bytes:
        """Export data to PDF format using reportlab with table formatting.

        Args:
            data: List of dictionaries containing report data.
            title: Main title for the PDF document.
            subtitle: Optional subtitle for the document.

        Returns:
            PDF file content as bytes.

        Raises:
            ImportError: If reportlab is not installed.
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab not installed. Install with: pip install reportlab==4.0.0")

        if not data:
            return b""

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()

        # Title
        elements.append(Paragraph(title, styles['Heading1']))
        if subtitle:
            elements.append(Paragraph(subtitle, styles['Normal']))
        elements.append(Spacer(1, 20))

        # Timestamp
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 20))

        # Build table data
        headers = list(data[0].keys())
        table_data = [headers]
        for row in data:
            table_data.append([str(row.get(h, "")) for h in headers])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        doc.build(elements)

        output.seek(0)
        return output.getvalue()
